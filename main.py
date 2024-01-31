import tkinter as tk
from tkinter import Menu
from tkinter import ttk, filedialog, messagebox
from tkinter import *
from tkinter import Label
from PIL import Image, ImageTk
from tkinter import PhotoImage
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill


class MyGUI:

    def __init__(self):
        
        self.root = tk.Tk()
        self.root.iconbitmap('logo.ico')
        self.root.geometry('600x450')
        self.root.resizable(width=False, height=False)
        self.root.title('Checker invoice')
        self.root.protocol("WM_DELETE_WINDOW", self.confirm)
        
        self.custom_font = ('arial', 12)

        self.file_path = filedialog.askopenfilename(initialdir = "/", title = "Выберите реестор", filetypes=(("Книга Excel", "*.xlsx*"), ("С поддержкой макросов", "*.xlsm*"), ("Книга Excel 97-2003", "*.xls*")))
        self.sheet_name = "Сканированные"
        

        self.df = openpyxl.load_workbook(self.file_path, sheet_name=None)
        self.sheet_names = self.df.keys()
        # Call the function to get repeated values between two sheets
        self.repeated_values = self.get_repeated_values(self.sheet_names[1], self.sheet_names[2])
        # Call the function to color repeated values in the sheets to yellow
        self.color_repeated_values(self.sheet_names[1], self.repeated_values)
        self.color_repeated_values(self.sheet_names[2], self.repeated_values)

        #Frames (lable)
        #Left Frame---------------------------------------------------------------------
        self.leftFrame = ttk.Frame(self.root, width=300, height=300, relief=tk.GROOVE)
        self.leftFrame.pack_propagate(FALSE)
        self.leftFrame.grid(row=0, column=0)

        self.leftFrame_in = ttk.Frame(self.leftFrame)
        self.leftFrame_in.pack_propagate(TRUE)
        self.leftFrame_in.pack(anchor=tk.CENTER, pady=40 )
        
        #Right Frame---------------------------------------------------------------------
        self.rightFrame = ttk.Frame(self.root, width=300, height=300, relief=tk.GROOVE)
        self.rightFrame.pack_propagate(FALSE)
        self.rightFrame.grid(row=0, column=1)
        
        self.rightFrame_in = ttk.Frame(self.rightFrame)
        self.rightFrame_in.pack_propagate(TRUE)
        self.rightFrame_in.pack(anchor='center', expand=True)

        #Bottom left Frame---------------------------------------------------------------------
        self.bottomLeftFrame = ttk.Frame(self.root, width=300, height=150, relief=tk.RAISED)
        self.bottomLeftFrame.pack_propagate(FALSE)
        self.bottomLeftFrame.grid(row=1, column=0)

        self.bottomLeftFrame_in = ttk.Frame(self.bottomLeftFrame)
        self.bottomLeftFrame_in.pack_propagate(TRUE)
        self.bottomLeftFrame_in.pack(anchor=tk.CENTER, expand=True)
        
        #Bottom right Frame---------------------------------------------------------------------
        self.bottomRightFrame = ttk.Frame(self.root, width=300, height=150, relief=tk.RAISED)
        self.bottomRightFrame.pack_propagate(FALSE)
        self.bottomRightFrame.grid(row=1, column=1)

        self.bottomRightFrame_in = ttk.Frame(self.bottomRightFrame)
        self.bottomRightFrame_in.pack_propagate(TRUE)
        self.bottomRightFrame_in.pack(anchor=tk.CENTER, padx=10, pady=10)

        #Main Menu
        self.menuBar = tk.Menu(self.root)
        
        self.fileMenu = tk.Menu(self.menuBar, tearoff=0)
        self.fileMenu.add_command(label='Сохранить', command=self.save_data)
        self.fileMenu.add_command(label='О программе', command=self.SW_info)
        self.fileMenu.add_separator()
        self.fileMenu.add_command(label='Закрыть ПО', command=self.confirm)

        self.menuBar.add_cascade(label='Меню', menu=self.fileMenu)

        #Left System---------------------------------------------------------------------
        # Indicator of successful upload
        self.indicator = tk.Canvas(self.leftFrame_in, width=25, height=25)
        self.indicator.pack(expand=False, side=tk.TOP, pady=2)

        self.rectangle = self.indicator.create_rectangle(5, 5, 25, 25, fill="#FF3030", outline="black")
        
        # Button of upload a registry
        self.loadBtn = tk.Button(self.leftFrame_in, text='Загрузить реестор', font = self.custom_font, command=lambda: [self.execute_command(), self.load_data()])
        self.loadBtn.pack(expand=False, side=tk.TOP, pady=2)

        #break down line between drop down and button
        self.separator_frame1 = tk.Frame(self.leftFrame_in, width=200, height=2, bd=1, relief=tk.SUNKEN, bg="black")
        self.separator_frame1.pack(expand=FALSE, side=tk.TOP, pady=25)

        # text on the drop down menu
        self.label_nakl = tk.Label(self.leftFrame_in, text='Выберите столбец из списка ниже')
        self.label_nakl.pack(expand=False, side=tk.TOP, pady=2)

        # Drop Down Menu
        self.dropDown = ttk.Combobox(self.leftFrame_in, font=self.custom_font)
        self.dropDown.pack(expand=False, side=tk.TOP, pady=2)

        #Right System------------------------------------------------------------------------

        #Buttom for creating new sheet
        self.btn_new_sheet = tk.Button(self.rightFrame_in, text='Создать страницу', font=self.custom_font, command=self.create_sheet)
        self.btn_new_sheet.pack(side='top', pady=2, expand=False)

        #break down line between input and save
        self.separator_frame2 = tk.Frame(self.rightFrame_in, width=200, height=2, bd=1, relief=tk.SUNKEN, bg="black")
        self.separator_frame2.pack(expand=FALSE, side=tk.TOP, pady=25)

        #Input number of invoice
        self.input_text = tk.Label(self.rightFrame_in, text='Ввод номер наклодного через 2D сканер')
        self.input_text.pack(side='top', pady=2, expand=False)
        
        self.input_number = tk.Entry(self.rightFrame_in, font='14')
        self.input_number.pack(side='top', pady=2, expand=False)

        #Buttom for checking
        self.btn_check = tk.Button(self.rightFrame_in, text='Вводить значение', font=self.custom_font, command=self.checking)
        self.btn_check.pack(side='top', pady=2, expand=False)

        #Bottom right frame---------------------------------------------------------------------
        img = Image.open(r'C:\Users\mirdy\OneDrive\Рабочий стол\checker MSP\icon.png')
        self.resized_image= img.resize((250, 200))
        self.logo_image= ImageTk.PhotoImage(self.resized_image)
        
        self.img_lable = tk.Label(self.bottomRightFrame_in, image=self.logo_image)
        self.img_lable.pack(side='top', expand=False)

        #Bottom left frame---------------------------------------------------------------------

        #Save button
        self.save_button = tk.Button(self.bottomLeftFrame_in, text='Сохранить проверку', font=self.custom_font, command=self.save_data)
        self.save_button.pack(side='top', pady=2, expand=False)
        
        self.root.config(menu=self.menuBar)
        self.root.mainloop()

    def SW_info(self):
        messagebox.showinfo('О программе', 'ПО создал Е.Мирас в 2023 году')
    
    def confirm(self):
        answer = messagebox.askyesno(title='Confirmation', message='Вы уверены, что хотите закрыть?')
        if answer:
            self.root.destroy()

    def execute_command(self):
            if not self.file_path:
                self.indicator.itemconfig(self.rectangle, fill="#FF3030")
            else:
                self.indicator.itemconfig(self.rectangle, fill="#7FFF00")
        
    '''def saveProg(self):
        filedialog.asksaveasfile(filetypes=(("Книга Excel", "*.xlsl*"), ("С поддержкой макросов", "*.xlsm*"), ("Книга Excel 97-2003", "*.xls*")))
        '''

    def load_data(self):
        #self.file_path = filedialog.askopenfilename(initialdir = "/", title = "Выберите реестор", filetypes=(("Книга Excel", "*.xlsx*"), ("С поддержкой макросов", "*.xlsm*"), ("Книга Excel 97-2003", "*.xls*")))

        if self.file_path:
            print("Выбранный Excel файл:", self.file_path)
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            row_number = 5

            options = []
            for row in sheet.iter_rows(min_row=row_number, values_only=True):
                cell_value = row[1]
                options.append(cell_value)

        self.dropDown['values'] = options

    def create_sheet(self):
        if self.file_path:
            wb2 = openpyxl.load_workbook(self.file_path)
            wb2.create_sheet(self.sheet_name)
            wb2.save(self.file_path)
        print('Создана страница в:', self.file_path)

    def color_repeated_values(self):
        # load the excel file using openpyxl
        wb = openpyxl.load_workbook(self.file_path)
        sheets = wb.active
    
        # iterate over the rows in the sheet
        for index, row in self.df[self.sheet_names].iterrows():
            # iterate over the cells in the row
            for i, cell in enumerate(row):
                # check if the cell value is in the list of repeated values
                if cell in self.repeated_values:
                    # color the cell to yellow
                    sheets[self.sheet_names].cell(row=index + 1, column=i + 1).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
        # save the changes to the excel file
        wb.save(self.file_path)

    def get_repeated_values(self, sheet1, sheet2):
        # Assuming df is the DataFrame read from the Excel file
        unique_values_sheet1 = self.df[sheet1].drop_duplicates()
        unique_values_sheet2 = self.df[sheet2].drop_duplicates()
        repeated_values = unique_values_sheet1[unique_values_sheet1.isin(unique_values_sheet2)].index
        return repeated_values

        '''filedialog.asksaveasfile(filetypes=(("Книга Excel", "*.xlsl*"), ("С поддержкой макросов", "*.xlsm*"), ("Книга Excel 97-2003", "*.xls*")))'''
MyGUI()
        